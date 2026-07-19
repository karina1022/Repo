import time
from datetime import datetime
from pathlib import Path

import serial
from serial.tools import list_ports
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BAUD_RATE = 115200
LOG_FILE = Path(__file__).with_name("tx_event_log.xlsx")
SHEET_NAME = "發送端紀錄"

# PC 端只傳送事件指令字元；
# 6-bit Payload 與 CRC-2 由 STM32 發送端自行產生。
VALID_COMMANDS = {
    "0": "SAFE",
    "1": "VRU Level 1",
    "2": "VRU Level 2",
    "3": "VRU Level 3",
    "4": "BRAKE Level 1",
    "5": "BRAKE Level 2",
    "6": "BRAKE Level 3",
    "E": "SYSTEM ERROR",
}


def show_ports() -> None:
    ports = list(list_ports.comports())

    if not ports:
        print("目前找不到 COM Port")
        return

    print("可用的 COM Port：")
    for port in ports:
        print(f"  {port.device}: {port.description}")


def show_commands() -> None:
    print("\n可用指令：")
    for command, description in VALID_COMMANDS.items():
        print(f"  {command} = {description}")

    print("  q = 離開")


def get_press_time() -> str:
    """取得目前時間，固定為 HH:MM:SS.mmm。"""
    return datetime.now().strftime("%H:%M:%S.%f")[:-3]


def create_log_workbook() -> None:
    """建立新的 Excel 紀錄檔。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME

    headers = [
        "發送序號",
        "事件代碼",
        "事件名稱",
        "按鍵時間",
    ]
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:D1"
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["D"].width = 18

    workbook.save(LOG_FILE)


def get_next_sequence() -> int:
    """讀取既有 Excel 紀錄，取得下一筆發送序號。"""
    if not LOG_FILE.exists():
        create_log_workbook()
        return 1

    try:
        workbook = load_workbook(LOG_FILE, read_only=True, data_only=True)
        sheet = workbook[SHEET_NAME]
        last_sequence = 0

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            value = row[0]
            try:
                last_sequence = max(last_sequence, int(value))
            except (TypeError, ValueError):
                continue

        workbook.close()
        return last_sequence + 1

    except (OSError, KeyError, PermissionError) as error:
        print(f"讀取 Excel 紀錄失敗，序號將從 1 開始：{error}")
        return 1


def save_send_record(
    sequence: int,
    command: str,
    event_name: str,
    press_time: str,
) -> None:
    """將發送序號、事件與按鍵時間寫入 Excel。"""
    if not LOG_FILE.exists():
        create_log_workbook()

    workbook = load_workbook(LOG_FILE)

    if SHEET_NAME not in workbook.sheetnames:
        sheet = workbook.create_sheet(SHEET_NAME)
        sheet.append(["發送序號", "事件代碼", "事件名稱", "按鍵時間"])
    else:
        sheet = workbook[SHEET_NAME]

    sheet.append([
        sequence,
        command,
        event_name,
        press_time,
    ])

    row = sheet.max_row

    # 時間刻意以文字儲存，避免 Excel 自動縮成 03:18.3。
    time_cell = sheet.cell(row=row, column=4)
    time_cell.number_format = "@"

    for column in range(1, 5):
        sheet.cell(row=row, column=column).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    workbook.save(LOG_FILE)
    workbook.close()


def main() -> None:
    show_ports()

    port_name = input(
        "\n請輸入 USB-UART 的 COM Port，例如 COM5："
    ).strip()

    try:
        with serial.Serial(
            port=port_name,
            baudrate=BAUD_RATE,
            timeout=0.1,
            write_timeout=1,
        ) as ser:

            # 等待 STM32 與 USB-UART 穩定
            time.sleep(2)

            # 清除程式剛啟動時殘留的資料
            ser.reset_input_buffer()
            ser.reset_output_buffer()

            print(
                f"\n已連線到 {port_name}, "
                f"baud = {BAUD_RATE}"
            )

            show_commands()

            sequence = get_next_sequence()
            print(f"發送紀錄將儲存至：{LOG_FILE}")
            print(f"下一筆發送序號：{sequence}")

            while True:
                command = input("\nPC send: ").strip()

                if command.lower() == "q":
                    break

                # 小寫 e 統一轉成大寫 E
                if command.lower() == "e":
                    command = "E"

                if command not in VALID_COMMANDS:
                    print("只能輸入 0、1、2、3、4、5、6、E 或 q")
                    continue

                # 記錄輸入事件並按下 Enter 的時間，精準到毫秒。
                press_time = get_press_time()
                print(
                    f"按鍵時間：{press_time}｜"
                    f"事件：{command} = {VALID_COMMANDS[command]}"
                )

                # 清除上一個事件可能留下的 STM32 回覆
                ser.reset_input_buffer()

                # 傳送單一 ASCII 指令給 STM32
                ser.write(command.encode("ascii"))
                ser.flush()

                try:
                    save_send_record(
                        sequence=sequence,
                        command=command,
                        event_name=VALID_COMMANDS[command],
                        press_time=press_time,
                    )
                except PermissionError:
                    print(
                        "Excel 檔案目前可能正在開啟，無法寫入。"
                        "請先關閉 tx_event_log.xlsx，再重新測試這一筆。"
                    )
                    continue
                except OSError as error:
                    print(f"Excel 儲存失敗：{error}")
                    continue

                print(
                    f"已送出：SEQ={sequence}｜"
                    f"{command} = {VALID_COMMANDS[command]}"
                )
                print(f"紀錄已儲存：{LOG_FILE.name}")

                sequence += 1

                print("STM32 reply:")

                # 最多等 2 秒，直到收到 TIME 開頭的量測結果
                deadline = time.monotonic() + 2.0
                received_time_result = False

                while time.monotonic() < deadline:
                    raw_line = ser.readline()

                    if not raw_line:
                        continue

                    line = raw_line.decode(
                        "utf-8",
                        errors="ignore",
                    ).strip()

                    if not line:
                        continue

                    print(line)

                    # STM32 完成 IR 與 LCD 量測後，
                    # 會傳回 TIME 開頭的資料
                    if line.startswith("TIME,"):
                        received_time_result = True
                        break

                if not received_time_result:
                    print(
                        "2 秒內沒有收到 TIME 回覆，"
                        "請檢查 STM32 程式、UART 接線或鮑率"
                    )

    except serial.SerialException as error:
        print(f"\n串口連線失敗：{error}")
        print(
            "請確認 COM Port、接線，以及 COM Port "
            "沒有被其他程式占用。"
        )


if __name__ == "__main__":
    main()
